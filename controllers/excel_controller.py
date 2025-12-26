# En controllers/excel_controller.py

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime
import os


class ExcelController:
    def __init__(self, cotizacion_controller, aiu_manager):
        self.cotizacion_controller = cotizacion_controller
        self.aiu_manager = aiu_manager

    def bordes_marco_con_interior(self, sheet, start_row, end_row, start_col=1, end_col=6):
        # Estilos
        grueso = Side(border_style="medium", color="000000")
        delgado = Side(border_style="thin", color="000000")

        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = sheet.cell(row=row, column=col)

                # Por defecto, bordes internos son delgados
                left = delgado
                right = delgado
                top = delgado
                bottom = delgado

                # Bordes exteriores → gruesos
                if col == start_col:  # primera columna
                    left = grueso
                if col == end_col:  # última columna
                    right = grueso
                if row == start_row:  # primera fila
                    top = grueso
                if row == end_row:  # última fila
                    bottom = grueso

                cell.border = Border(left=left, right=right, top=top, bottom=bottom)


    def aplicar_bordes_totales(self, sheet, start_row, end_row, color_hex):
        """Aplica bordes exteriores gruesos y bordes interiores delgados a un bloque de celdas."""
        thin = Side(border_style="thin", color="000000")
        medium = Side(border_style="medium", color="000000")

        fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

        for row in range(start_row, end_row + 1):
            for col in range(1, 7):  # Columnas A (1) a F (6)
                cell = sheet.cell(row=row, column=col)
                cell.fill = fill

                # Determinar bordes según posición
                left = medium if col == 1 else thin
                right = medium if col == 6 else thin
                top = medium if row == start_row else thin
                bottom = medium if row == end_row else thin

                cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    def aplicar_bordes_totales_con_marco_grueso(self, sheet, start_row, end_row, start_col=1, end_col=6,
                                                color_hex="D9EAD3"):
        """
        Aplica bordes delgados internos y un marco grueso alrededor de un bloque rectangular.
        - start_row, end_row: filas del bloque.
        - start_col, end_col: columnas del bloque (por defecto A:F → 1:6).
        """
        thin = Side(border_style="thin", color="000000")
        medium = Side(border_style="medium", color="000000")
        fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = sheet.cell(row=row, column=col)
                cell.fill = fill

                # Determinar bordes por posición (marco grueso)
                left = medium if col == start_col else thin
                right = medium if col == end_col else thin
                top = medium if row == start_row else thin
                bottom = medium if row == end_row else thin

                cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    def _insertar_subtotal_capitulo(self, sheet, row_num, start_row, chapter_num):
        """Inserta subtotal y devuelve la siguiente fila y la referencia de celda."""
        subtotal_font = Font(bold=True, italic=True, color="FFFFFF")
        subtotal_fill = PatternFill(start_color="008080", end_color="008080", fill_type="solid")
        # Mantener la alineación para que no se rompa el ajuste automático de la fila
        alignment = Alignment(horizontal="right", vertical="center")

        sheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=5)

        cell_label = sheet.cell(row=row_num, column=1)
        cell_label.value = f"SUBTOTAL CAPÍTULO {chapter_num}.0"
        cell_label.font = subtotal_font
        cell_label.fill = subtotal_fill
        cell_label.alignment = alignment

        cell_total = sheet.cell(row=row_num, column=6)
        cell_total.value = f"=SUM(F{start_row}:F{row_num - 1})"
        cell_total.font = subtotal_font
        cell_total.fill = subtotal_fill
        cell_total.alignment = alignment
        cell_total.number_format = '"$"#,##0.00'

        # Aplicar bordes para que combine con el resto
        self.bordes_marco_con_interior(sheet, row_num, row_num)

        celda_referencia = f"F{row_num}"
        return row_num + 1, celda_referencia


    def generate_excel(self, items, activities, tipo_persona, administracion, imprevistos, utilidad, iva_utilidad, nombre_cliente="", ruta_personalizada=""):

        """
        Genera un archivo Excel de cotización profesional, manejando capítulos,
        formato de celdas y lógica de AIU/IVA.
        """
        # 1. --- CONFIGURACIÓN INICIAL DEL LIBRO Y LA HOJA ---
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Cotización"

        # 2. --- DEFINICIÓN DE ESTILOS REUTILIZABLES ---
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="008080", end_color="008080", fill_type="solid")
        chapter_font = Font(bold=True, color="FFFFFF")
        chapter_fill = PatternFill(start_color="004D40", end_color="004D40", fill_type="solid")
        total_font = Font(bold=True)

        ### SOLUCIÓN AL PROBLEMA 2 y 3: DEFINIR ALINEACIONES ###
        # Alineación para la descripción: centrada verticalmente y con ajuste de texto
        description_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        # Alineación para el resto de columnas: centrada en ambas direcciones
        center_alignment = Alignment(horizontal="center", vertical="center")

        # 3. --- CONFIGURACIÓN DE LA HOJA (ANCHOS Y ENCABEZADOS) ---
        sheet.column_dimensions['A'].width = 5
        sheet.column_dimensions['B'].width = 60
        sheet.column_dimensions['C'].width = 12
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 20

        headers = ["Item", "Descripción", "Cantidad", "Unidad", "Precio Unitario", "Total"]
        for col_num, header_title in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header_title
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment

        # 4. --- PROCESAMIENTO DE LOS ÍTEMS (CAPÍTULOS Y ACTIVIDADES) ---
        row_num = 2
        chapter_counter = 0
        activity_counter = 0
        chapter_start_row = None
        capitulos_subtotales_celdas = []

        for i, item in enumerate(items):
            if item['type'] == 'chapter':
                # Antes de iniciar un nuevo capítulo, cerramos el anterior (si existe)
                if chapter_counter > 0 and activity_counter > 0:
                    row_num, celda_subtotal = self._insertar_subtotal_capitulo(sheet, row_num, chapter_start_row,chapter_counter)
                    capitulos_subtotales_celdas.append(celda_subtotal)

                chapter_counter += 1
                activity_counter = 0

                sheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
                cell = sheet.cell(row=row_num, column=1)
                cell.value = f"{chapter_counter}.0 {item['name'].upper()}"
                cell.font = chapter_font
                cell.fill = chapter_fill
                cell.alignment = center_alignment
                row_num += 1
                chapter_start_row = row_num

            elif item['type'] == 'activity':
                activity_counter += 1
                item_label = f"{chapter_counter}.{activity_counter}"

                # Llenado de datos
                sheet.cell(row=row_num, column=1).value = item_label
                sheet.cell(row=row_num, column=2).value = item['descripcion']
                sheet.cell(row=row_num, column=3).value = float(item['cantidad'])
                sheet.cell(row=row_num, column=4).value = item['unidad']
                sheet.cell(row=row_num, column=5).value = float(item['valor_unitario'])
                sheet.cell(row=row_num, column=6).value = f"=C{row_num}*E{row_num}"

                # Estilos y formatos
                for col in range(1, 7):
                    sheet.cell(row=row_num,column=col).alignment = description_alignment if col == 2 else center_alignment
                sheet.cell(row=row_num, column=5).number_format = '"$"#,##0.00'
                sheet.cell(row=row_num, column=6).number_format = '"$"#,##0.00'
                row_num += 1

        if chapter_counter > 0 and activity_counter > 0:
            row_num, celda_subtotal = self._insertar_subtotal_capitulo(sheet, row_num, chapter_start_row,chapter_counter)
            capitulos_subtotales_celdas.append(celda_subtotal)


        # 5. --- CÁLCULO DE SUBTOTAL, TOTALES Y AIU ---
        subtotal_row_num = row_num
        subtotal_cell_address = f"F{subtotal_row_num}"

        # Fila de Subtotal
        sheet.merge_cells(f"A{subtotal_row_num}:E{subtotal_row_num}")
        sheet[f"A{subtotal_row_num}"].value = "TOTAL COSTOS DIRECTOS"
        sheet[f"A{subtotal_row_num}"].font = header_font
        sheet[f"A{subtotal_row_num}"].fill = header_fill
        sheet[f"A{subtotal_row_num}"].alignment = Alignment(horizontal="right")

        if capitulos_subtotales_celdas:
            formula_subtotales = ",".join(capitulos_subtotales_celdas)
            sheet[subtotal_cell_address].value = f"=SUM({formula_subtotales})"
        else:
            sheet[subtotal_cell_address].value = 0

        sheet[subtotal_cell_address].number_format = '"$"#,##0.00'
        sheet[subtotal_cell_address].font = header_font
        sheet[subtotal_cell_address].fill = header_fill
        self.bordes_marco_con_interior(sheet, subtotal_row_num, subtotal_row_num)
        row_num += 1

        if tipo_persona.lower() == "juridica":
            # Lógica completa para AIU
            # Aplicar estilos a las filas de totales

            admin_row_num = row_num
            impr_row_num = row_num + 1
            util_row_num = row_num + 2
            iva_row_num = row_num + 3
            total_row_num = row_num + 4
            # Aplicar estilos y bordes a las filas de totales
            for i in range(admin_row_num, total_row_num + 1):
                sheet[f"A{i}"].alignment = Alignment(horizontal="right")
                sheet[f"F{i}"].number_format = '"$"#,##0.00'
                sheet[f"A{i}"].font = total_font
                sheet[f"F{i}"].font = total_font
            # Aplicar color y bordes gruesos por fuera
            self.aplicar_bordes_totales(sheet, admin_row_num, total_row_num, color_hex="D9EAD3")

            # Administración
            sheet.merge_cells(f"A{admin_row_num}:E{admin_row_num}");
            sheet[f"A{admin_row_num}"].value = f"ADMINISTRACIÓN ({administracion}%)"
            sheet[f"F{admin_row_num}"].value = f"={subtotal_cell_address}*({administracion}/100)"

            # Imprevistos
            sheet.merge_cells(f"A{impr_row_num}:E{impr_row_num}");
            sheet[f"A{impr_row_num}"].value = f"IMPREVISTOS ({imprevistos}%)"
            sheet[f"F{impr_row_num}"].value = f"={subtotal_cell_address}*({imprevistos}/100)"

            # Utilidad
            sheet.merge_cells(f"A{util_row_num}:E{util_row_num}");
            sheet[f"A{util_row_num}"].value = f"UTILIDAD ({utilidad}%)"
            sheet[f"F{util_row_num}"].value = f"={subtotal_cell_address}*({utilidad}/100)"

            # IVA sobre Utilidad
            sheet.merge_cells(f"A{iva_row_num}:E{iva_row_num}");
            sheet[f"A{iva_row_num}"].value = f"IVA SOBRE UTILIDAD ({iva_utilidad}%)"
            sheet[f"F{iva_row_num}"].value = f"=F{util_row_num}*({iva_utilidad}/100)"

            # Fila de Total Final
            sheet.merge_cells(f"A{total_row_num}:E{total_row_num}");
            sheet[f"A{total_row_num}"].value = "VALOR TOTAL COTIZACIÓN"
            sheet[f"F{total_row_num}"].value = f"=SUM(F{subtotal_row_num}:F{iva_row_num})"
            self.bordes_marco_con_interior(sheet,total_row_num,total_row_num)

            # Aplicar estilos a las filas de totales
            for i in range(admin_row_num, total_row_num + 1):
                sheet[f"A{i}"].alignment = Alignment(horizontal="right")
                sheet[f"F{i}"].number_format = '"$"#,##0.00'
                if i == total_row_num:
                    sheet[f"A{i}"].font = header_font
                    sheet[f"F{i}"].font = header_font
                    sheet[f"A{i}"].fill = header_fill
                    sheet[f"F{i}"].fill = header_fill
            self.bordes_marco_con_interior(sheet, 1, total_row_num)

        else:  # Persona Natural
            # Lógica simple con IVA sobre el subtotal
            iva_row_num = row_num
            total_row_num = row_num + 1

            # IVA
            sheet.merge_cells(f"A{iva_row_num}:E{iva_row_num}");
            sheet[f"A{iva_row_num}"].value = f"IVA ({iva_utilidad}%)"
            sheet[f"F{iva_row_num}"].value = f"={subtotal_cell_address}*({iva_utilidad}/100)"
            sheet[f"A{iva_row_num}"].alignment = Alignment(horizontal="right")
            sheet[f"F{iva_row_num}"].number_format = '"$"#,##0.00'

            # Fila de Total Final
            sheet.merge_cells(f"A{total_row_num}:E{total_row_num}");
            sheet[f"A{total_row_num}"].value = "VALOR TOTAL COTIZACIÓN"
            sheet[f"F{total_row_num}"].value = f"=SUM(F{subtotal_row_num}, F{iva_row_num})"
            sheet[f"A{total_row_num}"].font = header_font;
            sheet[f"F{total_row_num}"].font = header_font
            sheet[f"A{total_row_num}"].alignment = Alignment(horizontal="right")
            sheet[f"F{total_row_num}"].number_format = '"$"#,##0.00'
            sheet[f"A{total_row_num}"].fill = header_fill
            sheet[f"F{total_row_num}"].fill = header_fill
            self.bordes_marco_con_interior(sheet, total_row_num, total_row_num)

            self.aplicar_bordes_totales_con_marco_grueso(sheet, iva_row_num, iva_row_num, color_hex="D9EAD3")
            self.bordes_marco_con_interior(sheet, 1, total_row_num)

        # 6. --- GUARDAR EL ARCHIVO ---
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if ruta_personalizada and os.path.exists(ruta_personalizada):
            export_dir = ruta_personalizada

        else:
            # Fallback a la carpeta local si no hay ruta seleccionada
            export_dir = "exports"
            if not os.path.exists(export_dir):
                os.makedirs(export_dir)

        safe_name = "".join(c for c in nombre_cliente if c.isalnum() or c in (" ", "_", "-")).strip()
        safe_name = safe_name.replace(" ", "_")

        # Construir la ruta final
        filename = f"cotizacion_{safe_name}_{timestamp}.xlsx"
        excel_path = os.path.join(export_dir, filename)

        try:
            workbook.save(excel_path)
            print(f"Archivo Excel guardado en: {excel_path}")
            return excel_path
        except Exception as e:
            print(f"Error al guardar el archivo Excel: {e}")
            return None
